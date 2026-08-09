import os
import re
import time
import base64
import sys
import json
import csv
import sqlite3
from datetime import datetime
from datetime import timedelta
from flask import Flask, request, jsonify, render_template, session, send_from_directory, send_file
from PIL import Image, ImageOps, ImageFilter, ImageEnhance
from io import BytesIO
from sticker_packs import draw_sticker_pack, STICKER_PACK_OPTIONS

RESOURCE_DIR = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
APP_DATA_DIR = os.path.dirname(sys.executable) if getattr(sys, 'frozen', False) else RESOURCE_DIR
BUNDLED_STATIC_DIR = os.path.join(RESOURCE_DIR, 'static')
PHOTOS_DIR = os.path.join(APP_DATA_DIR, 'static', 'photos')
SETTINGS_PATH = os.path.join(APP_DATA_DIR, 'photobooth_settings.json')
TOKENS_DB_PATH = os.path.join(APP_DATA_DIR, 'photobooth_tokens.sqlite3')
DEFAULT_SETTINGS = {
    'session_duration_minutes': 4,
}
FRAME_WIDTH = 1182
# A modest footer below the fourth photo keeps the strip balanced without
# leaving an oversized blank area.
FRAME_HEIGHT = 3700


def load_settings():
    settings = DEFAULT_SETTINGS.copy()
    try:
        if os.path.exists(SETTINGS_PATH):
            with open(SETTINGS_PATH, 'r', encoding='utf-8') as settings_file:
                saved_settings = json.load(settings_file)
            if isinstance(saved_settings, dict):
                settings.update(saved_settings)
    except (OSError, json.JSONDecodeError):
        pass

    try:
        settings['session_duration_minutes'] = max(1, min(30, int(settings['session_duration_minutes'])))
    except (TypeError, ValueError):
        settings['session_duration_minutes'] = DEFAULT_SETTINGS['session_duration_minutes']
    return settings


def save_settings(settings):
    temporary_path = f'{SETTINGS_PATH}.tmp'
    with open(temporary_path, 'w', encoding='utf-8') as settings_file:
        json.dump(settings, settings_file, indent=2)
    os.replace(temporary_path, SETTINGS_PATH)


settings = load_settings()
if not os.path.exists(SETTINGS_PATH):
    save_settings(settings)

# PyInstaller extracts bundled files to a temporary folder for every run.
# Serve bundled web assets explicitly so uploaded photos can persist beside EXE.
app = Flask(__name__, static_folder=None)
# Secure secret key for session management – persists across restarts
# so that existing client session cookies remain valid.
_secret_path = os.path.join(APP_DATA_DIR, '.flask_secret')
if os.path.exists(_secret_path):
    with open(_secret_path, 'rb') as _f:
        app.secret_key = _f.read()
else:
    app.secret_key = os.urandom(24)
    with open(_secret_path, 'wb') as _f:
        _f.write(app.secret_key)
# Keep the server session alive for the configured kiosk duration plus a buffer.
app.permanent_session_lifetime = timedelta(minutes=settings['session_duration_minutes'] + 1)

# Ensure static/photos directory exists
os.makedirs(PHOTOS_DIR, exist_ok=True)


def get_token_db():
    connection = sqlite3.connect(TOKENS_DB_PATH)
    connection.row_factory = sqlite3.Row
    return connection


def initialise_token_db():
    with get_token_db() as connection:
        connection.execute('''
            CREATE TABLE IF NOT EXISTS tokens (
                token_number TEXT PRIMARY KEY,
                customer_name TEXT NOT NULL,
                contact_number TEXT DEFAULT '',
                email TEXT DEFAULT '',
                people_count INTEGER NOT NULL DEFAULT 1,
                amount REAL NOT NULL DEFAULT 0,
                payment_mode TEXT DEFAULT '',
                booth_used INTEGER NOT NULL DEFAULT 0,
                photo_given INTEGER NOT NULL DEFAULT 0,
                printing_done INTEGER NOT NULL DEFAULT 0,
                is_test INTEGER NOT NULL DEFAULT 0,
                created_at TEXT NOT NULL,
                booth_used_at TEXT,
                photo_given_at TEXT,
                printing_done_at TEXT
            )
        ''')
        connection.execute('''
            INSERT OR IGNORE INTO tokens
                (token_number, customer_name, amount, payment_mode, is_test, created_at)
            VALUES ('0', 'Test Customer', 0, 'Test', 1, ?)
        ''', (datetime.now().isoformat(timespec='seconds'),))
        columns = {row['name'] for row in connection.execute('PRAGMA table_info(tokens)')}
        if 'email' not in columns:
            connection.execute("ALTER TABLE tokens ADD COLUMN email TEXT DEFAULT ''")
        if 'people_count' not in columns:
            connection.execute('ALTER TABLE tokens ADD COLUMN people_count INTEGER NOT NULL DEFAULT 1')
        if 'printing_done' not in columns:
            connection.execute('ALTER TABLE tokens ADD COLUMN printing_done INTEGER NOT NULL DEFAULT 0')
        if 'printing_done_at' not in columns:
            connection.execute('ALTER TABLE tokens ADD COLUMN printing_done_at TEXT')


def normalise_token_number(value):
    return str(value or '').strip()


def token_as_dict(row):
    return dict(row) if row else None


def get_token(token_number):
    with get_token_db() as connection:
        row = connection.execute(
            'SELECT * FROM tokens WHERE token_number = ?', (normalise_token_number(token_number),)
        ).fetchone()
    return token_as_dict(row)


initialise_token_db()


@app.route('/static/photos/<path:filename>')
def serve_photo(filename):
    """Serve persistent photos saved beside this executable."""
    return send_from_directory(PHOTOS_DIR, filename)


@app.route('/api/photo/download/<path:filename>')
def download_photo(filename):
    """Force a local browser download for a saved photo or collage."""
    return send_from_directory(
        PHOTOS_DIR,
        filename,
        as_attachment=True,
        download_name=os.path.basename(filename),
    )


@app.route('/static/<path:filename>')
def serve_static_asset(filename):
    """Serve CSS, JavaScript, and other bundled assets."""
    return send_from_directory(BUNDLED_STATIC_DIR, filename)

def sanitize_filename(name: str) -> str:
    """Return a filesystem‑safe version of a user supplied name."""
    cleaned = re.sub(r'[^a-zA-Z0-9\s_-]', '', name).strip()
    return cleaned.replace(' ', '_')


def apply_photo_filter(image: Image.Image, filter_name: str) -> Image.Image:
    """Apply the same named filter choices available in the admin editor."""
    image = image.convert('RGB')
    filter_name = filter_name if filter_name in {
        'normal', 'grayscale', 'sepia', 'cyan', 'neon', 'contrast'
    } else 'normal'

    if filter_name == 'grayscale':
        return ImageOps.grayscale(image).convert('RGB')
    if filter_name == 'sepia':
        filtered = ImageOps.colorize(ImageOps.grayscale(image), '#372217', '#f1d4a0')
        filtered = ImageEnhance.Contrast(filtered).enhance(0.9)
        return ImageEnhance.Brightness(filtered).enhance(0.95)
    if filter_name in {'cyan', 'neon'}:
        hue_shift = 128 if filter_name == 'cyan' else 206
        hue, saturation, value = image.convert('HSV').split()
        hue = hue.point(lambda pixel: (pixel + hue_shift) % 256)
        filtered = Image.merge('HSV', (hue, saturation, value)).convert('RGB')
        return ImageEnhance.Color(filtered).enhance(1.1 if filter_name == 'cyan' else 1.4)
    if filter_name == 'contrast':
        filtered = ImageEnhance.Contrast(image).enhance(1.4)
        return ImageEnhance.Brightness(filtered).enhance(1.05)
    return image

@app.route('/')
def kiosk():
    return render_template('kiosk.html')

@app.route('/admin')
def admin():
    return render_template('admin.html', sticker_packs=STICKER_PACK_OPTIONS)


@app.route('/api/settings')
def get_settings():
    """Return the kiosk-safe portion of the venue configuration."""
    return jsonify({'session_duration_minutes': settings['session_duration_minutes']})


@app.route('/api/admin/settings', methods=['POST'])
def update_settings():
    """Persist venue timer settings."""
    data = request.json or {}

    try:
        duration = int(data.get('session_duration_minutes'))
    except (TypeError, ValueError):
        return jsonify({'error': 'Enter a whole number of minutes'}), 400
    if not 1 <= duration <= 30:
        return jsonify({'error': 'Session duration must be between 1 and 30 minutes'}), 400

    settings['session_duration_minutes'] = duration
    settings.pop('admin_pin', None)
    save_settings(settings)
    app.permanent_session_lifetime = timedelta(minutes=duration + 1)
    return jsonify({'status': 'success', 'session_duration_minutes': duration})


def token_analytics(connection):
    rows = connection.execute('SELECT * FROM tokens WHERE is_test = 0').fetchall()
    tokens = [dict(row) for row in rows]
    return {
        'total_customers': len(tokens),
        'total_people': sum(max(1, int(token['people_count'] or 1)) for token in tokens),
        'total_revenue': sum(float(token['amount'] or 0) for token in tokens),
        'pending_booth': sum(not token['booth_used'] for token in tokens),
        'booth_used': sum(bool(token['booth_used']) for token in tokens),
        'pending_prints': sum(bool(token['booth_used']) and not token['printing_done'] for token in tokens),
        'printing_done': sum(bool(token['printing_done']) for token in tokens),
        'photos_given': sum(bool(token['photo_given']) for token in tokens),
    }


@app.route('/api/admin/tokens', methods=['GET'])
def list_tokens():
    with get_token_db() as connection:
        rows = connection.execute(
            'SELECT * FROM tokens ORDER BY is_test ASC, CAST(token_number AS INTEGER), token_number'
        ).fetchall()
        return jsonify({'tokens': [dict(row) for row in rows], 'analytics': token_analytics(connection)})


@app.route('/api/admin/tokens/next', methods=['GET'])
def next_token_number():
    """Reserve no record, but provide the next numeric token for the entry form."""
    with get_token_db() as connection:
        rows = connection.execute('SELECT token_number FROM tokens').fetchall()
    numeric_tokens = [int(row['token_number']) for row in rows if str(row['token_number']).isdigit()]
    return jsonify({'token_number': str((max(numeric_tokens) if numeric_tokens else 0) + 1)})


@app.route('/api/admin/tokens', methods=['POST'])
def save_token():
    data = request.json or {}
    token_number = normalise_token_number(data.get('token_number'))
    customer_name = str(data.get('customer_name') or '').strip()
    if not token_number or not customer_name:
        return jsonify({'error': 'Token number and customer name are required'}), 400
    try:
        amount = float(data.get('amount') or 0)
    except (TypeError, ValueError):
        return jsonify({'error': 'Amount must be a valid number'}), 400
    try:
        people_count = max(1, int(data.get('people_count') or 1))
    except (TypeError, ValueError):
        return jsonify({'error': 'Number of people must be a whole number'}), 400
    now = datetime.now().isoformat(timespec='seconds')
    with get_token_db() as connection:
        connection.execute('''
            INSERT INTO tokens (token_number, customer_name, contact_number, email, people_count, amount, payment_mode, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(token_number) DO UPDATE SET
              customer_name = excluded.customer_name,
              contact_number = excluded.contact_number,
              email = excluded.email,
              people_count = excluded.people_count,
              amount = excluded.amount,
              payment_mode = excluded.payment_mode
        ''', (token_number, customer_name, str(data.get('contact_number') or '').strip(),
              str(data.get('email') or '').strip(), people_count, amount, str(data.get('payment_mode') or '').strip(), now))
    return jsonify({'status': 'success'})


@app.route('/api/admin/tokens/<token_number>', methods=['DELETE'])
def delete_token(token_number):
    if token_number == '0':
        return jsonify({'error': 'The permanent test token cannot be deleted'}), 400
    with get_token_db() as connection:
        token = connection.execute('SELECT * FROM tokens WHERE token_number = ?', (token_number,)).fetchone()
        if not token:
            return jsonify({'error': 'Customer token not found'}), 404
            
        token_dict = dict(token)
        session_dir = sanitize_filename(token_number)
        folder_path = os.path.join(PHOTOS_DIR, session_dir)
        if os.path.exists(folder_path) and os.path.isdir(folder_path):
            archive_dir = os.path.join(PHOTOS_DIR, 'archived')
            os.makedirs(archive_dir, exist_ok=True)
            cust_name = sanitize_filename(token_dict.get('customer_name', 'Unknown'))
            created_at = token_dict.get('created_at', '')
            date_str = created_at.split('T')[0] if 'T' in created_at else created_at.split(' ')[0]
            if not date_str:
                date_str = datetime.now().strftime('%Y-%m-%d')
            archive_name = f"{cust_name}_{token_number}_{date_str}"
            archive_path = os.path.join(archive_dir, archive_name)
            if os.path.exists(archive_path):
                archive_path = f"{archive_path}_{int(time.time())}"
            import shutil
            shutil.move(folder_path, archive_path)

        connection.execute('DELETE FROM tokens WHERE token_number = ?', (token_number,))
    return jsonify({'status': 'success'})

@app.route('/api/admin/tokens', methods=['DELETE'])
def delete_all_tokens():
    with get_token_db() as connection:
        tokens = connection.execute('SELECT * FROM tokens WHERE token_number != "0"').fetchall()
        import shutil
        archive_dir = os.path.join(PHOTOS_DIR, 'archived')
        os.makedirs(archive_dir, exist_ok=True)
        
        for token in tokens:
            token_dict = dict(token)
            token_number = token_dict['token_number']
            session_dir = sanitize_filename(token_number)
            folder_path = os.path.join(PHOTOS_DIR, session_dir)
            if os.path.exists(folder_path) and os.path.isdir(folder_path):
                cust_name = sanitize_filename(token_dict.get('customer_name', 'Unknown'))
                created_at = token_dict.get('created_at', '')
                date_str = created_at.split('T')[0] if 'T' in created_at else created_at.split(' ')[0]
                if not date_str:
                    date_str = datetime.now().strftime('%Y-%m-%d')
                archive_name = f"{cust_name}_{token_number}_{date_str}"
                archive_path = os.path.join(archive_dir, archive_name)
                if os.path.exists(archive_path):
                    archive_path = f"{archive_path}_{int(time.time())}"
                try:
                    shutil.move(folder_path, archive_path)
                except Exception as e:
                    print(f"Error archiving {folder_path}: {e}")
                    
        connection.execute('DELETE FROM tokens WHERE token_number != "0"')
        
    return jsonify({'status': 'success', 'deleted': len(tokens)})


@app.route('/api/admin/tokens/<token_number>/status', methods=['POST'])
def update_token_status(token_number):
    data = request.json or {}
    field = data.get('field')
    if field not in {'booth_used', 'printing_done', 'photo_given'}:
        return jsonify({'error': 'Invalid status field'}), 400
    value = 1 if data.get('value') else 0
    timestamp_field = f'{field}_at'
    with get_token_db() as connection:
        result = connection.execute(
            f'UPDATE tokens SET {field} = ?, {timestamp_field} = ? WHERE token_number = ?',
            (value, datetime.now().isoformat(timespec='seconds') if value else None, token_number),
        )
        if not result.rowcount:
            return jsonify({'error': 'Token not found'}), 404
    return jsonify({'status': 'success'})


@app.route('/api/admin/tokens/export')
def export_tokens():
    try:
        from openpyxl import Workbook
    except ImportError:
        return jsonify({'error': 'Excel export requires openpyxl. Rebuild after installing requirements.txt.'}), 500
    with get_token_db() as connection:
        rows = [dict(row) for row in connection.execute('SELECT * FROM tokens ORDER BY token_number')]
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Customers'
    headers = ['Token No', 'Name', 'Contact No', 'Email', 'No. of People', 'Amount', 'Payment Mode', 'Used Booth', 'Printing Done', 'Photo Given', 'Created At', 'Booth Used At', 'Printing Done At', 'Photo Given At']
    sheet.append(headers)
    for row in rows:
        sheet.append([row['token_number'], row['customer_name'], row['contact_number'], row['email'], row['people_count'], row['amount'], row['payment_mode'],
                      'Yes' if row['booth_used'] else 'No', 'Yes' if row['printing_done'] else 'No', 'Yes' if row['photo_given'] else 'No', row['created_at'],
                      row['booth_used_at'] or '', row['printing_done_at'] or '', row['photo_given_at'] or ''])
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(24, max(12, max(len(str(cell.value or '')) for cell in column) + 2))
    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return send_file(stream, as_attachment=True, download_name='photobooth_customers.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@app.route('/api/admin/tokens/import', methods=['POST'])
def import_tokens():
    upload = request.files.get('file')
    if not upload or not upload.filename:
        return jsonify({'error': 'Choose an Excel or CSV file first'}), 400
    try:
        if upload.filename.lower().endswith('.csv'):
            rows = list(csv.DictReader(upload.stream.read().decode('utf-8-sig').splitlines()))
        else:
            from openpyxl import load_workbook
            workbook = load_workbook(upload, read_only=True, data_only=True)
            sheet = workbook.active
            values = list(sheet.iter_rows(values_only=True))
            headers = [str(value or '').strip() for value in values[0]]
            rows = [dict(zip(headers, row)) for row in values[1:] if any(value is not None for value in row)]
    except Exception as error:
        return jsonify({'error': f'Could not read file: {error}'}), 400

    def find(row, *names):
        lowered = {str(key).strip().lower(): value for key, value in row.items()}
        for name in names:
            if name.lower() in lowered:
                return lowered[name.lower()]
        return ''

    imported = 0
    with get_token_db() as connection:
        for row in rows:
            token_number = normalise_token_number(find(row, 'Token No', 'Token Number', 'Token'))
            customer_name = str(find(row, 'Name', 'Customer Name') or '').strip()
            if not token_number or not customer_name:
                continue
            try:
                amount = float(find(row, 'Amount') or 0)
            except (TypeError, ValueError):
                amount = 0
            try:
                people_count = max(1, int(find(row, 'No. of People', 'No of People', 'People', 'People Count') or 1))
            except (TypeError, ValueError):
                people_count = 1
            connection.execute('''
                INSERT INTO tokens (token_number, customer_name, contact_number, email, people_count, amount, payment_mode, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(token_number) DO UPDATE SET customer_name=excluded.customer_name,
                  contact_number=excluded.contact_number, email=excluded.email, people_count=excluded.people_count, amount=excluded.amount, payment_mode=excluded.payment_mode
            ''', (token_number, customer_name, str(find(row, 'Contact No', 'Contact Number', 'Phone') or ''),
                  str(find(row, 'Email', 'Email Address') or ''), people_count, amount, str(find(row, 'Payment Mode', 'Payment') or ''),
                  datetime.now().isoformat(timespec='seconds')))
            imported += 1
    return jsonify({'status': 'success', 'imported': imported})

# ---------------------------------------------------------------------------
# Session handling – the kiosk start endpoint also acts as a login
# ---------------------------------------------------------------------------
@app.before_request
def make_session_permanent():
    session.permanent = True

@app.route('/api/token/login', methods=['POST'])
def token_login():
    """Validate a customer token and record that the booth has been used."""
    data = request.json or {}
    token_number = normalise_token_number(data.get('token_number'))
    if not token_number:
        return jsonify({'error': 'Enter your token number'}), 400

    token = get_token(token_number)
    if not token:
        return jsonify({'error': 'This token was not found. Please ask a staff member for help.'}), 404
    if token['booth_used'] and not token['is_test']:
        return jsonify({'error': 'This token has already been used. Please ask a staff member for help.'}), 409

    now = datetime.now().isoformat(timespec='seconds')
    with get_token_db() as connection:
        connection.execute(
            'UPDATE tokens SET booth_used = 1, booth_used_at = COALESCE(booth_used_at, ?) WHERE token_number = ?',
            (now, token_number),
        )

    session['token_number'] = token_number
    session['customer_name'] = token['customer_name']
    session['session_dir'] = sanitize_filename(token_number) or 'token'
    os.makedirs(os.path.join(PHOTOS_DIR, session['session_dir']), exist_ok=True)
    return jsonify({
        'token_number': token_number,
        'customer_name': token['customer_name'],
        'session_dir': session['session_dir'],
        'is_test': bool(token['is_test']),
    })

@app.route('/api/session/start', methods=['POST'])
def start_session():
    data = request.json or {}
    token_number = normalise_token_number(data.get('token_number') or session.get('token_number'))
    token = get_token(token_number)
    if not token:
        return jsonify({'error': 'A valid token is required to start a session'}), 401
    customer_name = token['customer_name']
    session_dir_name = sanitize_filename(token_number) or 'token'
    session_path = os.path.join(PHOTOS_DIR, session_dir_name)
    os.makedirs(session_path, exist_ok=True)
    # Store in Flask session for later requests
    session['customer_name'] = customer_name
    session['token_number'] = token_number
    session['session_dir'] = session_dir_name
    return jsonify({
        'session_dir': session_dir_name,
        'customer_name': customer_name,
        'token_number': token_number,
    })

# ---------------------------------------------------------------------------
# Photo upload – accepts images, stores each with a unique timestamp suffix, builds collage
# ---------------------------------------------------------------------------
@app.route('/api/session/upload', methods=['POST'])
def upload_photos():
    data = request.json or {}
    # Prefer the session‑stored directory if not explicitly supplied
    session_dir = data.get('session_dir') or session.get('session_dir')
    image_data_list = data.get('images', [])
    image_filter = data.get('image_filter', 'normal')
    if not session_dir or not image_data_list:
        return jsonify({'error': 'Missing session_dir or images'}), 400
    session_path = os.path.join(PHOTOS_DIR, session_dir)
    if not os.path.exists(session_path):
        return jsonify({'error': 'Session directory not found'}), 404
    saved_files = []
    pil_images = []
    session_timestamp = str(int(time.time()))
    for idx, img_base64 in enumerate(image_data_list):
        try:
            if ',' in img_base64:
                img_base64 = img_base64.split(',')[1]
            img_bytes = base64.b64decode(img_base64)
            img = Image.open(BytesIO(img_bytes)).convert('RGB')
            # Apply an Unsharp Mask to significantly increase photo sharpness
            img = img.filter(ImageFilter.UnsharpMask(radius=1.5, percent=100, threshold=3))
            
            # Save individual photo – timestamp ensures uniqueness within user folder
            filename = f"capture_{session_timestamp}_{idx + 1}.jpg"
            filepath = os.path.join(session_path, filename)
            img.save(filepath, 'JPEG', quality=100, subsampling=0)
            saved_files.append(f"/static/photos/{session_dir}/{filename}")
            pil_images.append(img)
        except Exception as e:
            return jsonify({'error': f"Failed to process image {idx + 1}: {str(e)}"}), 500
    # Build vertical collage
    collage_url = None
    if pil_images:
        try:
            # Dimensions for 600 DPI (doubled from 300 DPI for higher resolution)
            # 50mm x 157mm frame with a subtle footer below the last photo.
            frame_w, frame_h = FRAME_WIDTH, FRAME_HEIGHT
            # 43.3mm x 31.8mm photo
            photo_w, photo_h = 1022, 752
            # 300 DPI alternative:
            # frame_w, frame_h = 591, 1772
            # photo_w, photo_h = 511, 376
            
            # Margins
            left_margin = (frame_w - photo_w) // 2
            top_margin = left_margin
            gutter = 80
            
            frame_color_hex = data.get('frame_color', '#ffffff')
            try:
                bg_color = tuple(int(frame_color_hex.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
            except Exception:
                bg_color = (255, 255, 255)
            
            collage = Image.new('RGB', (frame_w, frame_h), bg_color)
            current_y = top_margin
            
            for img in pil_images:
                # Crop/resize photo to exact dimensions
                scaled_img = ImageOps.fit(apply_photo_filter(img, image_filter), (photo_w, photo_h), Image.Resampling.LANCZOS)
                collage.paste(scaled_img, (left_margin, current_y))
                current_y += photo_h + gutter
            
            # Apply sticker pack if selected
            sticker_pack = data.get('sticker_pack', 'none')
            if sticker_pack and sticker_pack != 'none':
                draw_sticker_pack(collage, sticker_pack)
                
            collage_filename = f"collage_{session_timestamp}.jpg"
            collage_filepath = os.path.join(session_path, collage_filename)
            collage.save(collage_filepath, 'JPEG', quality=100, subsampling=0)
            collage_url = f"/static/photos/{session_dir}/{collage_filename}"
        except Exception as e:
            return jsonify({'error': f"Failed to build collage: {str(e)}"}), 500
    return jsonify({
        'status': 'success',
        'files': saved_files,
        'collage_url': collage_url,
        'session_timestamp': session_timestamp
    })

@app.route('/api/session/render_preview', methods=['POST'])
def render_preview():
    data = request.json or {}
    session_dir = data.get('session_dir') or session.get('session_dir')
    image_data_list = data.get('images', [])
    session_timestamp = data.get('session_timestamp')
    frame_color_hex = data.get('frame_color', '#ffffff')
    image_filter = data.get('image_filter', 'normal')
    layer_type = data.get('layer_type', 'full')
    
    if not session_dir:
        return jsonify({'error': 'Missing session_dir'}), 400
        
    if not image_data_list and not session_timestamp:
        return jsonify({'error': 'Missing images and session_timestamp'}), 400

    try:
        frame_w, frame_h = FRAME_WIDTH, FRAME_HEIGHT
        photo_w, photo_h = 1022, 752
        # 300 DPI alternative:
        # frame_w, frame_h = 591, 1772
        # photo_w, photo_h = 511, 376
        left_margin = (frame_w - photo_w) // 2
        top_margin = left_margin
        gutter = 80

        if layer_type == 'stickers_only':
            collage = Image.new('RGBA', (frame_w, frame_h), (0, 0, 0, 0))
            sticker_pack = data.get('sticker_pack', 'none')
            if sticker_pack and sticker_pack != 'none':
                draw_sticker_pack(collage, sticker_pack)
            buffer = BytesIO()
            collage.save(buffer, 'PNG')
            buffer.seek(0)
            base64_str = base64.b64encode(buffer.getvalue()).decode('utf-8')
            return jsonify({
                'status': 'success',
                'preview_data': f"data:image/png;base64,{base64_str}"
            })

        try:
            bg_color = tuple(int(frame_color_hex.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        except Exception:
            bg_color = (255, 255, 255)

        collage = Image.new('RGB', (frame_w, frame_h), bg_color)
        current_y = top_margin
        
        # Load from base64 list OR from local files
        pil_images = []
        if image_data_list:
            for img_base64 in image_data_list:
                if ',' in img_base64:
                    img_base64 = img_base64.split(',')[1]
                img_bytes = base64.b64decode(img_base64)
                img = Image.open(BytesIO(img_bytes)).convert('RGB')
                pil_images.append(img)
        else:
            session_path = os.path.join(PHOTOS_DIR, session_dir)
            for i in range(1, 5):
                filepath = os.path.join(session_path, f"capture_{session_timestamp}_{i}.jpg")
                if os.path.exists(filepath):
                    img = Image.open(filepath).convert('RGB')
                    pil_images.append(img)
                    
        if not pil_images:
            return jsonify({'error': 'No images found to preview'}), 400

        for img in pil_images:
            # Apply unsharp mask to match capture flow, only if loaded from disk, but let's do it for all to be consistent
            img = img.filter(ImageFilter.UnsharpMask(radius=1.5, percent=100, threshold=3))
            scaled_img = ImageOps.fit(apply_photo_filter(img, image_filter), (photo_w, photo_h), Image.Resampling.LANCZOS)
            collage.paste(scaled_img, (left_margin, current_y))
            current_y += photo_h + gutter

        # Apply sticker pack if selected
        sticker_pack = data.get('sticker_pack', 'none')
        if sticker_pack and sticker_pack != 'none':
            draw_sticker_pack(collage, sticker_pack)

        buffer = BytesIO()
        collage.save(buffer, 'JPEG', quality=85)
        buffer.seek(0)
        base64_str = base64.b64encode(buffer.getvalue()).decode('utf-8')
        return jsonify({
            'status': 'success',
            'preview_data': f"data:image/jpeg;base64,{base64_str}"
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@app.route('/api/session/edit_existing', methods=['POST'])
def edit_existing():
    """Rebuilds the collage for an existing session with new frame color and stickers."""
    data = request.json or {}
    session_dir = data.get('session_dir')
    session_timestamp = data.get('session_timestamp')
    frame_color_hex = data.get('frame_color', '#ffffff')
    sticker_pack = data.get('sticker_pack', 'none')
    image_filter = data.get('image_filter', 'normal')
    
    if not session_dir or not session_timestamp:
        return jsonify({'error': 'Missing session_dir or session_timestamp'}), 400
        
    session_path = os.path.join(PHOTOS_DIR, session_dir)
    if not os.path.exists(session_path):
        return jsonify({'error': 'Session directory not found'}), 404
        
    try:
        frame_w, frame_h = FRAME_WIDTH, FRAME_HEIGHT
        photo_w, photo_h = 1022, 752
        # 300 DPI alternative:
        # frame_w, frame_h = 591, 1772
        # photo_w, photo_h = 511, 376
        left_margin = (frame_w - photo_w) // 2
        top_margin = left_margin
        gutter = 80
        
        try:
            bg_color = tuple(int(frame_color_hex.lstrip('#')[i:i+2], 16) for i in (0, 2, 4))
        except Exception:
            bg_color = (255, 255, 255)
            
        collage = Image.new('RGB', (frame_w, frame_h), bg_color)
        current_y = top_margin
        # Generate a new timestamp to uniquely identify this edit
        edit_timestamp = str(int(time.time()))
        
        images_found = False
        for i in range(1, 5):
            old_filepath = os.path.join(session_path, f"capture_{session_timestamp}_{i}.jpg")
            if os.path.exists(old_filepath):
                images_found = True
                img = Image.open(old_filepath).convert('RGB')
                scaled_img = ImageOps.fit(apply_photo_filter(img, image_filter), (photo_w, photo_h), Image.Resampling.LANCZOS)
                collage.paste(scaled_img, (left_margin, current_y))
                current_y += photo_h + gutter
                
        if not images_found:
            return jsonify({'error': 'No original capture images found'}), 404
            
        if sticker_pack and sticker_pack != 'none':
            draw_sticker_pack(collage, sticker_pack)
            
        # Save as an edited collage under the same session
        collage_filename = f"collage_edited_{session_timestamp}_{edit_timestamp}.jpg"
        collage_filepath = os.path.join(session_path, collage_filename)
        collage.save(collage_filepath, 'JPEG', quality=100, subsampling=0)
        collage_url = f"/static/photos/{session_dir}/{collage_filename}"
        
        return jsonify({
            'status': 'success',
            'collage_url': collage_url
        })
    except Exception as e:
        return jsonify({'error': f"Failed to edit collage: {str(e)}"}), 500


# ---------------------------------------------------------------------------
# Admin – list all sessions
# ---------------------------------------------------------------------------
@app.route('/api/admin/sessions', methods=['GET'])
def get_sessions():
    try:
        sessions = []
        with get_token_db() as connection:
            token_names = {
                row['token_number']: row['customer_name']
                for row in connection.execute('SELECT token_number, customer_name FROM tokens')
            }
        if os.path.exists(PHOTOS_DIR):
            for folder_name in sorted(os.listdir(PHOTOS_DIR)):
                folder_path = os.path.join(PHOTOS_DIR, folder_name)
                if not os.path.isdir(folder_path): continue
                customer_name = token_names.get(folder_name, folder_name.replace('_', ' '))
                
                sessions_dict = {}
                for file_name in os.listdir(folder_path):
                    if not (file_name.endswith('.jpg') or file_name.endswith('.gif')):
                        continue
                    
                    parts = file_name.split('_')
                    if len(parts) >= 2:
                        if file_name.startswith('capture_'):
                            ts = parts[1]
                        elif file_name.startswith('collage_edited_'):
                            # collage_edited_{session_ts}.jpg  OR  collage_edited_{session_ts}_{edit_ts}.jpg
                            ts = parts[2].split('.')[0]
                        elif file_name.startswith('collage_'):
                            ts = parts[1].split('.')[0]
                        elif file_name.startswith('animation_'):
                            ts = parts[1].split('.')[0]
                        else:
                            continue
                            
                        if ts not in sessions_dict:
                            try:
                                formatted_time = time.ctime(int(ts))
                            except:
                                formatted_time = ts
                            sessions_dict[ts] = {
                                'folder': folder_name,
                                'customer_name': customer_name,
                                'timestamp': ts,
                                'time': formatted_time,
                                'files': [],
                                'collage_url': None,
                                'collage_edited_urls': [],
                                'gif_url': None
                            }
                        
                        file_path = f"/static/photos/{folder_name}/{file_name}"
                        if file_name.startswith('capture_'):
                            sessions_dict[ts]['files'].append(file_path)
                        elif file_name.startswith('collage_edited_'):
                            sessions_dict[ts]['collage_edited_urls'].append(file_path)
                        elif file_name.startswith('collage_'):
                            sessions_dict[ts]['collage_url'] = file_path
                        elif file_name.startswith('animation_'):
                            sessions_dict[ts]['gif_url'] = file_path
                
                for ts in sessions_dict:
                    sess = sessions_dict[ts]
                    sess['files'].sort()
                    sess['collage_edited_urls'].sort()
                    sessions.append(sess)
        
        sessions.sort(key=lambda x: x['timestamp'], reverse=True)
        return jsonify({'sessions': sessions})
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------------------------------------------------------------------------
# Customer – retrieve only their own sessions
# ---------------------------------------------------------------------------
@app.route('/api/customer/gallery', methods=['GET'])
def customer_gallery():
    customer_name = session.get('customer_name')
    if not customer_name:
        return jsonify({'error': 'Not logged in'}), 401
    sessions = []
    session_dir = session.get('session_dir')
    if not session_dir:
        return jsonify({'error': 'No active token session'}), 401
    folder_path = os.path.join(PHOTOS_DIR, session_dir)
    
    if os.path.exists(folder_path) and os.path.isdir(folder_path):
        sessions_dict = {}
        for file_name in os.listdir(folder_path):
            if not (file_name.endswith('.jpg') or file_name.endswith('.gif')):
                continue
                
            parts = file_name.split('_')
            if len(parts) >= 2:
                if file_name.startswith('capture_'):
                    ts = parts[1]
                elif file_name.startswith('collage_edited_'):
                    ts = parts[2].split('.')[0]
                elif file_name.startswith('collage_'):
                    ts = parts[1].split('.')[0]
                elif file_name.startswith('animation_'):
                    ts = parts[1].split('.')[0]
                else:
                    continue
                    
                if ts not in sessions_dict:
                    sessions_dict[ts] = {
                        'folder': session_dir,
                        'customer_name': customer_name,
                        'timestamp': ts,
                        'files': [],
                        'collage_url': None,
                        'collage_edited_urls': [],
                        'gif_url': None
                    }
                
                file_path = f"/static/photos/{session_dir}/{file_name}"
                if file_name.startswith('capture_'):
                    sessions_dict[ts]['files'].append(file_path)
                elif file_name.startswith('collage_edited_'):
                    sessions_dict[ts]['collage_edited_urls'].append(file_path)
                elif file_name.startswith('collage_'):
                    sessions_dict[ts]['collage_url'] = file_path
                elif file_name.startswith('animation_'):
                    sessions_dict[ts]['gif_url'] = file_path
        
        for ts in sorted(sessions_dict.keys(), reverse=True):
            sess = sessions_dict[ts]
            sess['files'].sort()
            sess['collage_edited_urls'].sort()
            sessions.append(sess)
            
    return jsonify({'sessions': sessions})

# ---------------------------------------------------------------------------
# Delete photo
# ---------------------------------------------------------------------------
@app.route('/api/customer/photo', methods=['DELETE'])
def delete_photo():
    customer_name = session.get('customer_name')
    if not customer_name:
        return jsonify({'error': 'Not logged in'}), 401
    
    data = request.json or {}
    file_url = data.get('file_url')
    if not file_url:
        return jsonify({'error': 'Missing file_url'}), 400
        
    sanitized_name = sanitize_filename(customer_name)
    # Ensure the user can only delete their own photos
    expected_prefix = f"/static/photos/{sanitized_name}/"
    if not file_url.startswith(expected_prefix):
        return jsonify({'error': 'Unauthorized to delete this file'}), 403
        
    filename = file_url[len(expected_prefix):]
    # Prevent directory traversal
    if '/' in filename or '\\' in filename:
        return jsonify({'error': 'Invalid filename'}), 400
        
    file_path = os.path.join(PHOTOS_DIR, sanitized_name, filename)
    if os.path.exists(file_path):
        try:
            os.remove(file_path)
            return jsonify({'status': 'success'})
        except Exception as e:
            return jsonify({'error': f'Failed to delete file: {str(e)}'}), 500
    else:
        return jsonify({'error': 'File not found'}), 404

# ---------------------------------------------------------------------------
# Admin edit save
# ---------------------------------------------------------------------------
@app.route('/api/admin/save_edit', methods=['POST'])
def save_edit():
    data = request.json or {}
    session_dir = data.get('session_dir')
    timestamp = data.get('timestamp')
    edited_img_base64 = data.get('image')
    if not session_dir or not edited_img_base64:
        return jsonify({'error': 'Missing session_dir or image data'}), 400
    session_path = os.path.join(PHOTOS_DIR, session_dir)
    if not os.path.exists(session_path):
        return jsonify({'error': 'Session directory not found'}), 404
    try:
        if ',' in edited_img_base64:
            edited_img_base64 = edited_img_base64.split(',')[1]
        img_bytes = base64.b64decode(edited_img_base64)
        img = Image.open(BytesIO(img_bytes)).convert('RGB')
        # Use current time as unique edit identifier to support multiple edits
        import time as time_mod
        edit_ts = str(int(time_mod.time()))
        filename = f"collage_edited_{timestamp}_{edit_ts}.jpg" if timestamp else f"collage_edited_{edit_ts}.jpg"
        filepath = os.path.join(session_path, filename)
        img.save(filepath, 'JPEG', quality=95)
        return jsonify({
            'status': 'success',
            'collage_edited_url': f"/static/photos/{session_dir}/{filename}"
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ---------------------------------------------------------------------------
# Logout
# ---------------------------------------------------------------------------
@app.route('/api/customer/logout', methods=['POST'])
def logout():
    session.clear()
    return jsonify({'status': 'logged out'})

if __name__ == '__main__':
    use_ssl = '--ssl' in sys.argv
    ssl_ctx = 'adhoc' if use_ssl else None
    if use_ssl:
        print("Starting Flask with adhoc SSL (HTTPS) enabled.")
    # The Flask debug reloader relaunches a PyInstaller EXE recursively and
    # can prevent the packaged server from binding to its port.
    app.run(host='0.0.0.0', port=5000, debug=False, use_reloader=False, ssl_context=ssl_ctx)
